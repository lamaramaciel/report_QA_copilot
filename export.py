"""Create the Excel review workbook for Slide QA Copilot."""
from __future__ import annotations

import io
import json
from collections import Counter
from datetime import date, datetime
from numbers import Number
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import IllegalCharacterError
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

EXPORT_BUILD = "2026.08.02.2"
EXCEL_CELL_TEXT_LIMIT = 32767

DARK = PatternFill("solid", fgColor="1F2937")
TEAL = PatternFill("solid", fgColor="0F766E")
WHITE = Font(color="FFFFFF", bold=True)
WRAP = Alignment(vertical="top", wrap_text=True)
THIN = Side(style="thin", color="E5E7EB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _valid_xml_10_character(character: str) -> bool:
    """Return whether one Unicode character is legal in an Excel XML cell.

    This is intentionally stricter than removing only the usual C0 controls. It
    also excludes lone surrogate code points and U+FFFE/U+FFFF, which may enter
    extracted PDF text or model verbatims and can break workbook generation.
    """
    codepoint = ord(character)
    return (
        codepoint in (0x09, 0x0A, 0x0D)
        or 0x20 <= codepoint <= 0xD7FF
        or 0xE000 <= codepoint <= 0xFFFD
        or 0x10000 <= codepoint <= 0x10FFFF
    )


def _clean_text(value: str) -> str:
    cleaned = "".join(ch for ch in value if _valid_xml_10_character(ch))
    return cleaned[:EXCEL_CELL_TEXT_LIMIT]


def _clean(value: Any) -> Any:
    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass

    # Preserve native scalar values where Excel handles them safely.
    if isinstance(value, (bool, Number, datetime, date)):
        return value

    if isinstance(value, bytes):
        value = value.decode("utf-8", errors="replace")
    elif isinstance(value, (dict, list, tuple, set)):
        value = json.dumps(value, ensure_ascii=False, default=str)
    elif not isinstance(value, str):
        value = str(value)

    return _clean_text(value)


def _write_cell(ws, row: int, column: int, value: Any):
    """Write a value with a final defensive fallback.

    A malformed source string should never prevent the analyst from downloading
    the rest of the review workbook.
    """
    cleaned = _clean(value)
    try:
        return ws.cell(row=row, column=column, value=cleaned)
    except IllegalCharacterError:
        fallback = _clean_text(str(cleaned).encode("utf-8", "replace").decode("utf-8", "replace"))
        return ws.cell(row=row, column=column, value=fallback)


def _write_dataframe(ws, df: pd.DataFrame, widths: dict[str, int] | None = None) -> None:
    for col_idx, col in enumerate(df.columns, start=1):
        c = _write_cell(ws, 1, col_idx, col)
        c.fill = DARK
        c.font = WHITE
        c.border = BORDER
        c.alignment = WRAP

    for r_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
        for c_idx, value in enumerate(row, start=1):
            cleaned = _clean(value)
            c = _write_cell(ws, r_idx, c_idx, cleaned)
            c.border = BORDER
            c.alignment = WRAP
            if isinstance(cleaned, str) and cleaned.lower().startswith(("http://", "https://")):
                # Hyperlink targets are sanitized separately as well.
                c.hyperlink = _clean_text(cleaned)
                c.font = Font(color="1D4ED8", underline="single")

    for idx, col in enumerate(df.columns, start=1):
        width = (widths or {}).get(str(col), 22)
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"
    if len(df.columns):
        ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{max(1, len(df) + 1)}"


def _sanitize_entire_workbook(wb: Workbook) -> None:
    """Final pass over every string, including summary cells and headers."""
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    cell.value = _clean_text(cell.value)
                if cell.hyperlink and getattr(cell.hyperlink, "target", None):
                    cell.hyperlink.target = _clean_text(str(cell.hyperlink.target))


def build_report(results: list[dict]) -> bytes:
    claims_rows: list[dict] = []
    source_rows: list[dict] = []
    usage_rows: list[dict] = []

    for result in results:
        claims_rows.append({k: v for k, v in result.items() if k not in {"sources", "usage"}})
        for source in result.get("sources", []) or []:
            source_rows.append({
                "Claim ID": result.get("Claim ID", ""),
                "Slide": result.get("Slide", ""),
                "Slide Title": result.get("Slide Title", ""),
                **source,
            })
        usage = result.get("usage") or {}
        usage_rows.append({
            "Claim ID": result.get("Claim ID", ""),
            "Slide": result.get("Slide", ""),
            **usage,
        })

    claims_df = pd.DataFrame(claims_rows)
    sources_df = pd.DataFrame(source_rows)
    usage_df = pd.DataFrame(usage_rows)

    wb = Workbook()
    summary = wb.active
    summary.title = "QA Summary"
    _write_cell(summary, 1, 1, "Slide QA Summary")
    summary["A1"].font = Font(size=16, bold=True)

    status_counts = Counter(claims_df.get("Status", pd.Series(dtype=str)).tolist())
    metrics = [
        ("Claims checked", len(claims_df)),
        ("Confirmed", status_counts.get("✅ Confirmed", 0)),
        ("Partial", status_counts.get("⚠️ Partial", 0)),
        ("Incorrect", status_counts.get("❌ Incorrect", 0)),
        ("Not Found", status_counts.get("❓ Not Found", 0)),
        ("Inaccessible", status_counts.get("🔒 Inaccessible", 0)),
        ("No Reference", status_counts.get("❓ No Reference", 0)),
    ]

    for c_idx, name in enumerate(["Metric", "Value"], start=1):
        c = _write_cell(summary, 3, c_idx, name)
        c.fill = TEAL
        c.font = WHITE
        c.border = BORDER

    for r_idx, (name, value) in enumerate(metrics, start=4):
        _write_cell(summary, r_idx, 1, name).border = BORDER
        _write_cell(summary, r_idx, 2, value).border = BORDER

    total_cost = (
        sum(float(v or 0) for v in usage_df.get("estimated_cost_usd", pd.Series(dtype=float)).tolist())
        if not usage_df.empty else 0
    )
    total_tokens = (
        sum(int(v or 0) for v in usage_df.get("total_tokens", pd.Series(dtype=int)).tolist())
        if not usage_df.empty else 0
    )

    _write_cell(summary, 13, 1, "Usage estimate").font = Font(bold=True, size=12)
    _write_cell(summary, 14, 1, "Total tokens reported")
    _write_cell(summary, 14, 2, total_tokens)
    _write_cell(summary, 15, 1, "Estimated model cost (USD)")
    _write_cell(summary, 15, 2, round(total_cost, 6))
    _write_cell(summary, 16, 1, "Note")
    _write_cell(summary, 16, 2, "Estimate only; verify against the provider billing console.")
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 58

    ws_claims = wb.create_sheet("QA Results")
    _write_dataframe(ws_claims, claims_df, {
        "Claim": 55,
        "Verdict": 60,
        "Supported Claims": 50,
        "Unsupported Claims": 50,
        "Unverified Claims": 50,
        "Direct Verbatims": 60,
        "Evidence by Source": 60,
        "Suggested Fix": 60,
        "URLs": 55,
    })

    ws_sources = wb.create_sheet("QA Sources")
    _write_dataframe(ws_sources, sources_df, {"url": 65, "Slide Title": 45, "raw_status": 32})

    ws_usage = wb.create_sheet("API Usage")
    _write_dataframe(ws_usage, usage_df)

    _sanitize_entire_workbook(wb)

    output = io.BytesIO()
    wb.save(output)

    # Validate that the generated bytes can be opened before returning them.
    output.seek(0)
    load_workbook(output, read_only=True).close()
    return output.getvalue()
